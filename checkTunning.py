import pandas as pd
import subprocess
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

# 读取Excel文件
df = pd.read_excel('requirements.xlsx')

# 定义一个函数，用于通过ADB获取实际值
def get_actual_value(command):
    try:
        result = subprocess.check_output(command, shell=True)
        return result.decode('utf-8').strip()
    except Exception as e:
        print(f"Error executing command {command}: {e}")
        return None

# 定义每个字段对应的ADB命令
commands = {
    'Android_version': 'adb shell getprop ro.build.version.release',
    'Kernel_version': 'adb shell "uname -r"',
    'User_version': 'adb shell getprop ro.build.type',
    'Memory': 'adb shell "cat /proc/meminfo | grep MemTotal"',
    'lmk_minfree_levels': 'adb shell "cat /proc/sys/vm/lmk_minfree_levels"',
    'dirty_ratio': 'adb shell "cat /proc/sys/vm/dirty_ratio"',
    'dirty_background_ratio': 'adb shell "cat /proc/sys/vm/dirty_background_ratio"',
    'readahead': 'adb shell "cat /sys/block/sda/queue/read_ahead_kb"',
    'io_scheduler': 'adb shell "cat /sys/block/sda/queue/scheduler"',
    'CPU_kernel_count': 'adb shell "cat /proc/cpuinfo | grep \'processor\' | wc -l"',
    'CPU_load_1min': 'adb shell "cat /proc/loadavg | awk \'{print $1}\'"',
    'CPU_load_5min': 'adb shell "cat /proc/loadavg | awk \'{print $2}\'"',
    'CPU_load_15min': 'adb shell "cat /proc/loadavg | awk \'{print $3}\'"',
    'Temperature': 'adb shell "cat /sys/class/thermal/thermal_zone0/temp"'
}

# 获取当前时间，用于命名输出文件
current_time = datetime.now().strftime('%Y%m%d%H%M%S')
output_file = f'results_{current_time}.xlsx'

# 遍历每一行，获取实际值并进行比较
for index, row in df.iterrows():
    key = row['mode']
    command = commands.get(key)
    
    if command:
        actual_value = get_actual_value(command)
        
        # 特殊处理Kernel_version
        if key == 'Kernel_version' and actual_value:
            actual_value = actual_value.split('-')[0]
        # 特殊处理Memory
        elif key == 'Memory' and actual_value:
            actual_value = str(int(actual_value.split()[1]) // 1024)
        elif key == 'Temperature' and actual_value:
            actual_value = float(actual_value) / 1000
            expected_value = float(row['requirements'])
            satisfy = 1 if actual_value < expected_value else 0
            df.at[index, 'actual'] = str(actual_value)
            df.at[index, 'satisfy'] = satisfy
            continue  # 跳过下面的通用比较
        
        expected_value = str(row['requirements'])
        satisfy = 1 if expected_value == actual_value else 0
        
        # 将实际值和比较结果添加到DataFrame
        df.at[index, 'actual'] = actual_value
        df.at[index, 'satisfy'] = satisfy
    else:
        df.at[index, 'actual'] = 'N/A'
        df.at[index, 'satisfy'] = 0

# 保存结果到新的Excel文件
df.to_excel(output_file, index=False)

# 使用openpyxl来优化Excel文件
wb = load_workbook(output_file)
ws = wb.active

# 设置列宽
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width

# 设置单元格样式
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
left_alignment = Alignment(horizontal="left")

for row in ws.iter_rows(min_row=2):  # 跳过标题行
    for cell in row:
        cell.alignment = left_alignment
    if row[-1].value == 0:  # 最后一列是satisfy列
        row[-1].fill = red_fill  # 只标记satisfy列

# 保存优化后的Excel文件
wb.save(output_file)

print(f"Optimized results saved to {output_file}")